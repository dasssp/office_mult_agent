import asyncio
import csv
import hashlib
import io
import json
from dataclasses import asdict, dataclass, field
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import ClassVar
from uuid import uuid4

from app.repositories.files import FileRepository
from app.schemas import RequestContext


class UnsafeFileError(ValueError):
    pass


@dataclass(frozen=True)
class FileMetadata:
    file_id: str
    filename: str
    content_type: str | None
    byte_size: int
    sha256: str
    row_count: int
    file_type: str
    storage_ref: str
    tenant_id: str
    created_by: str
    status: str = "stored"

    def as_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass
class FileService:
    files: dict[str, list[dict[str, object]]] = field(default_factory=dict)
    metadata: dict[str, FileMetadata] = field(default_factory=dict)
    storage_dir: Path = field(default_factory=lambda: Path("uploads"))
    max_bytes: int = 2 * 1024 * 1024
    max_rows: int = 10_000
    repository: FileRepository | None = None

    _allowed_types: ClassVar[set[str]] = {"csv", "json", "xlsx", "docx", "pdf"}
    _content_types: ClassVar[dict[str, set[str]]] = {
        "csv": {"text/csv", "application/csv"},
        "json": {"application/json", "text/json"},
        "xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        "docx": {"application/vnd.openxmlformats-officedocument.wordprocessingml.document"},
        "pdf": {"application/pdf"},
    }

    async def store_and_parse(
        self,
        *,
        filename: str,
        content: bytes,
        context: RequestContext,
        content_type: str | None = None,
    ) -> str:
        suffix = self._validate_upload(filename, content, content_type)
        rows = self._parse(suffix, content)
        self._validate_row_count(rows)
        file_id = str(uuid4())
        tenant_dir = self.storage_dir / hashlib.sha256(
            context.tenant_id.encode("utf-8")
        ).hexdigest()[:24]
        storage_path = tenant_dir / f"{file_id}.{suffix}"
        await asyncio.to_thread(tenant_dir.mkdir, parents=True, exist_ok=True)
        await asyncio.to_thread(storage_path.write_bytes, content)
        self.files[file_id] = rows
        metadata = FileMetadata(
            file_id=file_id,
            filename=filename,
            content_type=content_type,
            byte_size=len(content),
            sha256=hashlib.sha256(content).hexdigest(),
            row_count=len(rows),
            file_type=suffix,
            storage_ref=str(storage_path),
            tenant_id=context.tenant_id,
            created_by=context.operator_id,
        )
        self.metadata[file_id] = metadata
        if self.repository is not None:
            await self.repository.save(metadata, context)
        return file_id

    async def get_rows(self, file_id: str, context: RequestContext) -> list[dict[str, object]]:
        metadata = await self.get_metadata(file_id, context)
        cached = self.files.get(file_id)
        if cached is not None:
            return cached
        content = await asyncio.to_thread(Path(metadata.storage_ref).read_bytes)
        rows = self._parse(metadata.file_type, content)
        self._validate_row_count(rows)
        self.files[file_id] = rows
        return rows

    async def get_metadata(self, file_id: str, context: RequestContext) -> FileMetadata:
        metadata = self.metadata.get(file_id)
        if metadata is None and self.repository is not None:
            stored = await self.repository.get(file_id, context)
            if stored is not None:
                metadata = FileMetadata(**stored)  # type: ignore[arg-type]
                self.metadata[file_id] = metadata
        if metadata is None or metadata.tenant_id != context.tenant_id:
            raise KeyError(file_id)
        return metadata

    async def delete(self, file_id: str, context: RequestContext) -> None:
        metadata = await self.get_metadata(file_id, context)
        path = Path(metadata.storage_ref)
        if self.storage_dir.resolve() not in path.resolve().parents:
            raise UnsafeFileError("invalid storage reference")
        await asyncio.to_thread(path.unlink, missing_ok=True)
        self.files.pop(file_id, None)
        self.metadata.pop(file_id, None)
        if self.repository is not None:
            await self.repository.delete(file_id, context)

    def _validate_upload(self, filename: str, content: bytes, content_type: str | None) -> str:
        if not filename or Path(filename).name != filename or len(filename) > 255:
            raise UnsafeFileError("invalid filename")
        if any(ord(char) < 32 for char in filename):
            raise UnsafeFileError("invalid filename")
        if not content:
            raise UnsafeFileError("file is empty")
        if len(content) > self.max_bytes:
            raise UnsafeFileError("file exceeds maximum size")
        suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if suffix not in self._allowed_types:
            raise UnsafeFileError("unsupported file type")
        if content_type and content_type not in self._content_types[suffix]:
            raise UnsafeFileError("content type does not match filename")
        if suffix == "pdf" and not content.startswith(b"%PDF-"):
            raise UnsafeFileError("invalid PDF signature")
        if suffix in {"xlsx", "docx"} and not content.startswith(b"PK"):
            raise UnsafeFileError("invalid Office document signature")
        return suffix

    def _parse(self, suffix: str, content: bytes) -> list[dict[str, object]]:
        if suffix == "xlsx":
            return self._parse_xlsx(content)
        if suffix == "docx":
            return [{"text": self._parse_docx(content)}]
        if suffix == "pdf":
            return [{"text": self._parse_pdf(content)}]
        text = content.decode("utf-8-sig")
        return self._parse_csv(text) if suffix == "csv" else self._parse_json(text)

    def _validate_row_count(self, rows: list[dict[str, object]]) -> None:
        if len(rows) > self.max_rows:
            raise UnsafeFileError("file exceeds maximum row count")

    def _parse_csv(self, text: str) -> list[dict[str, object]]:
        rows = list(csv.DictReader(io.StringIO(text)))
        for row in rows:
            if any(
                isinstance(value, str) and value.startswith(("=", "+", "-", "@"))
                for value in row.values()
            ):
                raise UnsafeFileError("CSV formula injection detected")
        return [dict(row) for row in rows]

    def _parse_json(self, text: str) -> list[dict[str, object]]:
        parsed = json.loads(text)
        if not isinstance(parsed, list) or not all(isinstance(row, dict) for row in parsed):
            raise UnsafeFileError("JSON must be an array of objects")
        return parsed

    def _parse_xlsx(self, content: bytes) -> list[dict[str, object]]:
        from openpyxl import load_workbook

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
            handle.write(content)
            path = Path(handle.name)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
            values = list(workbook.active.iter_rows(values_only=True))
            workbook.close()
            if not values:
                return []
            headers = [str(value) if value is not None else "" for value in values[0]]
            if not all(headers) or len(set(headers)) != len(headers):
                raise UnsafeFileError("XLSX headers must be present and unique")
            return [dict(zip(headers, row, strict=True)) for row in values[1:]]
        finally:
            path.unlink(missing_ok=True)

    def _parse_docx(self, content: bytes) -> str:
        from docx import Document

        return "\n".join(paragraph.text for paragraph in Document(BytesIO(content)).paragraphs)

    def _parse_pdf(self, content: bytes) -> str:
        from pypdf import PdfReader

        return "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(content)).pages)
